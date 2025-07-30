import requests
import math
import openpyxl
from openpyxl.drawing.image import Image
import pandas as pd
from pathlib import Path
from datetime import date
from decimal import Decimal, ROUND_HALF_UP, getcontext
import random
import string



# 讀取所有價目表工作表
xlsx = pd.ExcelFile("ALL價目表.xlsx")
mini_df = pd.read_excel(xlsx, sheet_name='mini價目表')
prime_df = pd.read_excel(xlsx, sheet_name='prime價目表')
force_df = pd.read_excel(xlsx, sheet_name='force價目表')
df_cust = pd.read_excel(xlsx, sheet_name="客製價目表")
hupro_df = pd.read_excel(xlsx, sheet_name="製材所-琥珀木價目表")
wood_df = pd.read_excel(xlsx, sheet_name='木種設定')
wood_shape_df = pd.read_excel(xlsx, sheet_name='製材所形狀價格')
color_df = pd.read_excel(xlsx, sheet_name='桌板顏色價格')
shape_df = pd.read_excel(xlsx, sheet_name='桌板形狀價格')
foot_df = pd.read_excel(xlsx, sheet_name='桌腳價格')



def 查詢_mini(桌腳, 桌寬):
    row = mini_df[(mini_df['桌腳'] == 桌腳) & (mini_df['桌寬'] == 桌寬)]
    if row.empty:
        raise ValueError('桌板尺寸錯誤')
    else:
        return float(row['價格'].values[0]) 

def 查詢_prime(桌腳, 桌寬, 桌深):
    row = prime_df[
        (prime_df['桌腳'] == 桌腳) & 
        (prime_df['桌寬'] == 桌寬) & 
        (prime_df['桌深'] == 桌深)
    ]
    if row.empty:
        raise ValueError('桌板尺寸錯誤')
    else:
        return float(row['價格'].values[0]) 

def 查詢_force(桌寬, 桌深):
    row = force_df[(force_df['桌寬'] == 桌寬) & (force_df['桌深'] == 桌深)]
    if row.empty:
        raise ValueError('桌板尺寸錯誤')
    else:
        return float(row['價格'].values[0]) 

def 查詢客製桌板價格(桌寬, 桌深):
    for _, row in df_cust.iterrows():
        if row['桌寬(區間頭)'] <= 桌寬 <= row['桌寬(區間尾)'] and row['桌深(區間頭)'] <= 桌深 <= row['桌深(區間尾)']:
            return row['價格']
    if row.empty:
            raise ValueError('桌板尺寸錯誤')
    return None

def 查詢琥珀木價格(桌寬, 桌深):
    for _, row in hupro_df.iterrows():
        if row['桌寬(區間頭)'] <= 桌寬 <= row['桌寬(區間尾)'] and row['桌深(區間頭)'] <= 桌深 <= row['桌深(區間尾)']:
            return int(row['價格'])  # 去掉小數
    if row.empty:
            raise ValueError('桌板尺寸錯誤')
    return None
def 查詢木種設定(木種):
    row = wood_df[wood_df['木種'] == 木種]
    if row.empty:
            raise ValueError('木種錯誤')
    return float(row['價格乘積'].values[0]) 

def 查詢製材所形狀價格(製材所形狀):
    row = wood_shape_df[wood_shape_df['製材所形狀'] == 製材所形狀]
    if row.empty:
            raise ValueError('製材所形狀錯誤')
    return float(row['價格'].values[0]) 

def 查詢桌板顏色價格(桌板顏色):
    if 桌板顏色.strip() == '':
        return 0  # 空白形狀，價格當成0，不跳錯
    row = color_df[color_df['桌板顏色'] == 桌板顏色]
    if row.empty:
        raise ValueError(桌板顏色)
    return float(row['價格'].values[0])

def 查詢桌板形狀價格(桌板形狀):
    if 桌板形狀.strip() == '':
        return 0  # 空白形狀，價格當成0，不跳錯
    row = shape_df[shape_df['桌板形狀'] == 桌板形狀]
    if row.empty:
        raise ValueError(桌板形狀)
    return float(row['價格'].values[0]) 

def 查詢桌腳價格(桌腳):
    if 桌腳.strip() == '':
        return 0  # 空白桌腳，價格當成0，不跳錯
    row = foot_df[foot_df['桌腳'] == 桌腳]
    if row.empty:
        raise ValueError(桌腳)
    else:
        return float(row['價格'].values[0])

def 查詢單購桌板運費(桌板顏色):
    if 桌板顏色.strip() == '':
        return 200  #當纖維板運費200
    row = color_df[color_df['桌板顏色'] == 桌板顏色]
    if row.empty:
        raise ValueError(桌板顏色)
    else:
        return float(row['單購桌板運費'].values[0])

def 查詢製材所單購桌板運費(木種):
    row = wood_df[wood_df['木種'] == 木種]
    if row.empty:
        raise ValueError(木種)
    else:
        return float(row['單購桌板運費'].values[0])

getcontext().prec = 10  # 精度足夠即可

def excel_round(value, digits=-2):
    multiplier = Decimal('1e{}'.format(-digits))
    return int((Decimal(value) / multiplier).quantize(0, rounding=ROUND_HALF_UP) * multiplier)
         
while True:
    print('\n')
    try:
       報價 =input('請輸入規格或訂製或製材所或下單或報價單:')
    except Exception as e:
        print('輸入錯誤，原因:', e)
    else:
        if 報價 == '規格':
            try:
                桌寬 = float(input('請輸入桌寬:'))
                桌深 = float(input('請輸入桌深:'))
                桌腳 = input('請輸入桌腳:')
                桌板顏色 = input('請輸入桌板顏色:')
                桌板形狀 = input('請輸入桌板形狀:')

                桌腳price=查詢桌腳價格(桌腳)
                顏色price=查詢桌板顏色價格(桌板顏色)
                if 桌板形狀 == '':
                    桌板形狀 = '四方前上斜'
                形狀price=查詢桌板形狀價格(桌板形狀)
                             
            except Exception as e:
                print('輸入錯誤，原因:', e)
            else: 
                if 桌腳 == 'mini三節' or 桌腳 == 'mini二節' or 桌腳 == 'mini三節黑' or 桌腳 == 'mini三節白' or 桌腳 == 'mini二節黑' or 桌腳 == 'mini二節白':
                    try:
                        if 桌深 != 60:
                            print('mini規格桌深須為60!!!')
                            continue
                        mini升降桌price = 查詢_mini(桌腳, 桌寬)
                    except Exception as e:
                        print('輸入錯誤，原因:', e)
                    else:
                        total=mini升降桌price+顏色price+形狀price
        
                        print('和您報價')
                        print('%3.0f*%2.0f%s(%s)+%s=%5.0f'%(桌寬,桌深,桌板顏色,桌板形狀,桌腳,total))
                elif 桌腳 == 'prime三節' or 桌腳 == 'prime二節' or 桌腳 == 'prime三節黑' or 桌腳 == 'prime三節白' or 桌腳 == 'prime三節灰' or 桌腳 == 'prime二節黑' or 桌腳 == 'prime二節白' or 桌腳 =='prime二節灰':
                    try:       
                        prime升降桌price = 查詢_prime(桌腳, 桌寬, 桌深)
                    except Exception as e:
                        print('輸入錯誤，原因:', e)
                    else:
                        total=prime升降桌price+顏色price+形狀price
                        if 45<桌深<=60:
                            桌腳=桌腳+'(腳底座60公分)'
        
                        print('和您報價')
                        print('%3.0f*%2.0f%s(%s)+%s=%5.0f'%(桌寬,桌深,桌板顏色,桌板形狀,桌腳,total))
            
                elif 桌腳 == 'force' or 桌腳 =='force桌腳' or 桌腳 == 'force四柱桌腳' or 桌腳 == 'force四柱黑腳' or 桌腳 == 'force四柱白腳':
                    try:
                        force升降桌price = 查詢_force(桌寬, 桌深)
                    except Exception as e:
                        print('輸入錯誤，原因:', e)
                    else:
                        total=force升降桌price+顏色price+形狀price
                
                        print('和您報價')
                        if 顏色price != 0:
                            print('訂%3.0f*%2.0f*4%s(%s)+%s=%5.0f'%(桌寬,桌深,桌板顏色,桌板形狀,桌腳,total))
                        else:
                            print('%3.0f*%2.0f%s(%s)+%s=%5.0f'%(桌寬,桌深,桌板顏色,桌板形狀,桌腳,total))

        elif 報價 == '訂製':
            try:
                桌寬 = float(input('請輸入桌寬:'))
                桌深 = math.ceil(float(input('請輸入桌深:')))
                桌腳=input('請輸入桌腳:')
                桌板顏色=input('請輸入桌板顏色:')
                桌板形狀=input('請輸入桌板形狀:')
    
                price = 查詢客製桌板價格(桌寬, 桌深)
                桌腳price=查詢桌腳價格(桌腳)
                顏色price=查詢桌板顏色價格(桌板顏色)
                if 桌板形狀 == '':
                    桌板形狀 = '四方前上斜'
                形狀price=查詢桌板形狀價格(桌板形狀)
                運費price=查詢單購桌板運費(桌板顏色)
                
                def format_width_number(桌寬):
                    return str(int(桌寬)) if 桌寬 == int(桌寬) else str(桌寬)
                
            except Exception as e:
                print('輸入錯誤，原因:', e)
            else:
                total=price+桌腳price+顏色price+形狀price
                total=int(total)
                
                print('和您報價')
                if 桌腳price == 0 and 顏色price != 0 and 形狀price != 0:
                    total = total + 運費price
                    total = int(total)
                    print(f'單購桌板訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})={total}')
                    print('\n''備註:')
                    print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                    print('2. 訂製約45天(含假日)')
                    print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')                     
                elif 桌腳price == 0 and 顏色price != 0:
                    total = total + 運費price
                    total = int(total)
                    if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                        print('蜂巢板不能做上斜!!')
                        桌板形狀 = '四方全平'
                    print(f'單購桌板訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})={total}')
                    print('請自行判斷是否為訂製!!')
                    print('\n''備註:')
                    print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                    print('2. 訂製約35天(含假日)')
                    print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主') 
                elif 桌腳price == 0:
                    total = total + 運費price
                    total = int(total)
                    print(f'單購桌板訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})={total}')
                    print('請自行判斷是否為訂製!!')
                    print('\n''備註:')
                    print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                    print('2. 訂製約35天(含假日)')
                    print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主') 
                elif price == 0:
                    print('訂製尺寸有誤，注意桌深最小50')
                elif 桌寬<110 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰' or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳') and 桌深<57.5:
                    桌腳=桌腳+'(短版+短側片+腳底座45公分)'
                    if 顏色price != 0 and 形狀price != 0:
                        if 桌板形狀 == '四角導圓':
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主') 
                        else:
                            print('蜂巢板不可做凹!!')
                    elif 顏色price != 0:
                        if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                            print('蜂巢板不可做上斜!!')
                            桌板形狀 = '四方全平'
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                    else:
                        if 桌板形狀 == '四角導圓':
                            print('纖維板一定四角導圓，請刪除!!(否則報價有誤)')
                        else:
                            print(f'訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
        
                elif 桌寬<110 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime二節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰' or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳') and 桌深>=57.7:
                    if 57.5<=桌深<60:
                        桌腳=桌腳+'(短版+腳底座45公分)'
                        if 顏色price != 0 and 形狀price != 0:
                            if 桌板形狀 == '四角導圓':
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print('蜂巢板不可做凹!!')
                        elif 顏色price != 0:
                            if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                                print('蜂巢板不可做上斜!!')
                                桌板形狀 = '四方全平'
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            if 桌板形狀 == '四角導圓':
                                print('纖維板一定四角導圓，請刪除!!(否則報價有誤)')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                    elif 60<=桌深<68:
                        桌腳=桌腳+'(短版+腳底座60公分)'
                        if 顏色price != 0 and 形狀price != 0:
                            if 桌板形狀 == '四角導圓':
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print('蜂巢板不可做凹!!')
                        elif 顏色price != 0:
                            if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                                print('蜂巢板不可做上斜!!')
                                桌板形狀 = '四方全平'
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            if 桌板形狀 == '四角導圓':
                                print('纖維板一定四角導圓，請刪除!!(否則報價有誤)')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                    else:
                        桌腳=桌腳+'(短版)'
                        if 顏色price != 0 and 形狀price != 0:
                            if 桌板形狀 == '四角導圓':
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print('蜂巢板不可做凹!!')
                        elif 顏色price != 0:
                            if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                                print('蜂巢板不可做上斜!!')
                                桌板形狀 = '四方全平'
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            if 桌板形狀 == '四角導圓':
                                print('纖維板一定四角導圓，請刪除!!(否則報價有誤)')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
               
                elif 桌深<68 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime二節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰'or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳'):
                    if 57.5<=桌深<68:
                        桌腳=桌腳+'(腳底座60公分)'
                        if 顏色price != 0 and 形狀price != 0:
                            if 桌板形狀 == '四角導圓':
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print('蜂巢板不可做凹!!')
                        elif 顏色price != 0:
                            if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                                print('蜂巢板不可做上斜!!')
                                桌板形狀 = '四方全平'
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            if 桌板形狀 == '四角導圓':
                                print('纖維板一定四角導圓，請刪除!!(否則報價有誤)')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                    elif 桌深<57.5:
                        桌腳=桌腳+'(短側片+腳底座45公分)'
                        if 顏色price != 0 and 形狀price != 0:
                            if 桌板形狀 == '四角導圓':
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print('蜂巢板不可做凹!!')
                        elif 顏色price != 0:
                            if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                                print('蜂巢板不可做上斜!!')
                                桌板形狀 = '四方全平'
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            if 桌板形狀 == '四角導圓':
                                print('纖維板一定四角導圓，請刪除!!(否則報價有誤)')
                            else:
                                print(f'訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})+{桌腳}={total}')
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約35天(含假日)')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                elif (桌腳=='mini三節' or 桌腳=='mini二節' or 桌腳=='mini三節白' or 桌腳=='mini二節白' or 桌腳=='mini三節黑' or 桌腳=='mini二節黑'or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳'):               
                    if 顏色price != 0 and 形狀price != 0:
                        if 桌板形狀 == '四角導圓':
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            print('蜂巢板不可做凹!!')
                    elif 顏色price != 0:
                        if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                            print('蜂巢板不可做上斜!!')
                            桌板形狀 = '四方全平'
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                    else:
                        if 桌板形狀 == '四角導圓':
                            print('纖維板一定四角導圓，請刪除!!(否則報價有誤)')
                        else:
                            print(f'訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                elif 桌深<60 and (桌腳=='mini三節' or 桌腳=='mini二節' or 桌腳=='mini三節白' or 桌腳=='mini二節白' or 桌腳=='mini三節黑' or 桌腳=='mini二節黑'or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳'):
                    桌腳 = 桌腳+'(腳底座45公分)'
                    if 顏色price != 0 and 形狀price != 0:
                        if 桌板形狀 == '四角導圓':
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            print('蜂巢板不可做凹!!')
                    elif 顏色price != 0:
                        if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                            print('蜂巢板不可做上斜!!')
                            桌板形狀 = '四方全平'
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                    else:
                        if 桌板形狀 == '四角導圓':
                            print('纖維板一定四角導圓，請刪除!!(否則報價有誤)')
                        else:
                            print(f'訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                    
                elif 桌深<72 and (桌腳=='force' or 桌腳 == 'force桌腳' or 桌腳 == 'force四柱桌腳' or 桌腳 == 'force四柱白腳' or 桌腳 == 'force四柱黑腳'):
                    print('桌深無法安裝force!')
        
                else:
                    if 顏色price != 0 and 形狀price != 0:
                        if 桌板形狀 == '四角導圓':
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:
                            print('蜂巢板不可做凹!!')
                    elif 顏色price != 0:
                        if 桌板形狀 == '四方前上斜' or 桌板形狀 == '弧度上斜':
                            print('蜂巢板不可做上斜!!')
                            桌板形狀 = '四方全平'
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        else:    
                            print(f'訂{format_width_number(桌寬)}*{桌深}*4{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                    else:
                        if 桌板形狀 == '四角導圓':
                            print('纖維板一定四角導圓，請刪除!!(否則報價有誤)')
                        else:
                            print(f'訂{format_width_number(桌寬)}*{桌深}{桌板顏色}({桌板形狀})+{桌腳}={total}')
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約35天(含假日)')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
        elif 報價 == '製材所':
            try:
                木種=input('請輸入木種:')
            except Exception as e:
                print('輸入錯誤，原因:', e)
            else:
                if 木種 =='栓木脂接' or 木種 =='栓木直拼':
                    try:
                        桌寬=float(input('請輸入桌寬:'))
                        桌深=float(input('請輸入桌深:'))
                        厚度=float(input('請輸入厚度(2.7/3.5/4.5):'))
                        桌腳=input('請輸入桌腳:') 
                        製材所形狀=input('請輸入桌板形狀:')
                        桌腳price=查詢桌腳價格(桌腳)
                        運費price=查詢製材所單購桌板運費(木種)
                        
                        if 製材所形狀 =='':
                            製材所形狀 = '四方全平'
                        製材所形狀price=查詢製材所形狀價格(製材所形狀)
    
                        # ✅ 加入防呆：若輸入厚度為 3.3，自動修正為 3.5
                        if abs(厚度 - 3.3) < 0.05:
                            厚度 = 3.5
                            
                        # 這一行替換為 Decimal 計算（避免 float 誤差）
                        基礎價格 = Decimal(str(桌寬)) * Decimal(str(桌深)) * Decimal(str(厚度)) / Decimal('2700') * Decimal(str(查詢木種設定(木種))) 
                        桌板price = excel_round(基礎價格, -2)
                    except Exception as e:
                        print('輸入錯誤，原因:', e)
                    else:
                        if 桌腳 == '':
                            total=桌板price+製材所形狀price+運費price
                                
                            print('和您報價')
                            print('(製材所)-單購桌板訂%3.0f*%3.0f*%1.1f%s(%s)=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            
                        elif 桌深 >= 81 and 厚度<=2.7:
                                print('無法製作!!桌深厚度81cm以上，厚度需為3.3以上')
    
                        elif 桌寬<110 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰' or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳') and 桌深<57.5:
                            桌腳=桌腳+'(短版+短側片+腳底座45公分)'
                            total=桌板price+製材所形狀price+桌腳price
    
                            print('和您報價')
                            print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
        
                        elif 桌寬<110 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime二節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰' or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳') and 桌深>=57.7:
                            if 57.5<=桌深<60:
                                桌腳=桌腳+'(短版+腳底座45公分)'
                                total=桌板price+製材所形狀price+桌腳price
    
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                                    
                            elif 60<=桌深<68:
                                桌腳=桌腳+'(短版+腳底座60公分)'
                                total=桌板price+製材所形狀price+桌腳price
    
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                桌腳=桌腳+'(短版)'
                                total=桌板price+製材所形狀price+桌腳price
    
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        
               
                        elif 桌深<68 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime二節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰'or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳'):
                            if 57.5<=桌深<68:
                                桌腳=桌腳+'(腳底座60公分)'
                                total=桌板price+製材所形狀price+桌腳price
        
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            elif 桌深<57.5:
                                桌腳=桌腳+'(短側片+腳底座45公分)'
                                total=桌板price+製材所形狀price+桌腳price
        
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                                
                        elif 桌深<60 and (桌腳=='mini三節' or 桌腳=='mini二節' or 桌腳=='mini三節白' or 桌腳=='mini二節白' or 桌腳=='mini三節黑' or 桌腳=='mini二節黑'or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳'):
                            桌腳 = 桌腳+'(腳底座45公分)'
                            total=桌板price+製材所形狀price+桌腳price
        
                            print('和您報價')
                            print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            
                        elif 桌深<72 and (桌腳=='force' or 桌腳 == 'force桌腳' or 桌腳 == 'force四柱桌腳' or 桌腳 == 'force四柱白腳' or 桌腳 == 'force四柱黑腳'):
                            print('桌深無法安裝force!')
                        
                        else:
                            total=桌板price+製材所形狀price+桌腳price
    
                            print('和您報價')
                            print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            
                elif 木種 =='白橡木脂接' or 木種 =='白橡木直拼':
                    try:
                        桌寬=float(input('請輸入桌寬:'))
                        桌深=float(input('請輸入桌深:'))
                        厚度=float(input('請輸入厚度(2.7/3.3/4.5):'))
                        桌腳=input('請輸入桌腳:') 
                        製材所形狀=input('請輸入桌板形狀:')
                        桌腳price=查詢桌腳價格(桌腳)
                        運費price=查詢製材所單購桌板運費(木種)

                        if 製材所形狀 =='':
                            製材所形狀 = '四方全平'
                        製材所形狀price=查詢製材所形狀價格(製材所形狀)
                        
    
                        #✅ 加入防呆：若輸入厚度為 3.3，自動修正為 3.5
                        if abs(厚度 - 3.3) < 0.05:
                            厚度 = 3.5
                        # 這一行替換為 Decimal 計算（避免 float 誤差）
                        基礎價格 = Decimal(str(桌寬)) * Decimal(str(桌深)) * Decimal(str(厚度)) / Decimal('2700') * Decimal(str(查詢木種設定(木種))) 
                        桌板price = excel_round(基礎價格, -2)

                        #顯示厚度改回3.3
                        if 厚度 == 3.5:
                            厚度 = 3.3
                    except Exception as e:
                        print('輸入錯誤，原因:', e)
                    else:
                        if 桌腳 == '':
                            total=桌板price+製材所形狀price+運費price

                            print('和您報價')
                            print('(製材所)-單購桌板訂%3.0f*%3.0f*%1.1f%s(%s)=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            
                        elif 桌深 >= 81 and 厚度<=2.7:
                                print('無法製作!!桌深厚度81cm以上，厚度需為3.3以上')
    
                        elif 桌寬<110 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰' or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳') and 桌深<57.5:
                            桌腳=桌腳+'(短版+短側片+腳底座45公分)'
                            total=桌板price+製材所形狀price+桌腳price
    
                            print('和您報價')
                            print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
        
                        elif 桌寬<110 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime二節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰' or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳') and 桌深>=57.7:
                            if 57.5<=桌深<60:
                                桌腳=桌腳+'(短版+腳底座45公分)'
                                total=桌板price+製材所形狀price+桌腳price
    
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                                    
                            elif 60<=桌深<68:
                                桌腳=桌腳+'(短版+腳底座60公分)'
                                total=桌板price+製材所形狀price+桌腳price
    
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                桌腳=桌腳+'(短版)'
                                total=桌板price+製材所形狀price+桌腳price
    
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        
               
                        elif 桌深<68 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime二節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰'or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳'):
                            if 57.5<=桌深<68:
                                桌腳=桌腳+'(腳底座60公分)'
                                total=桌板price+製材所形狀price+桌腳price
        
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            elif 桌深<57.5:
                                桌腳=桌腳+'(短側片+腳底座45公分)'
                                total=桌板price+製材所形狀price+桌腳price
        
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                                
                        elif 桌深<60 and (桌腳=='mini三節' or 桌腳=='mini二節' or 桌腳=='mini三節白' or 桌腳=='mini二節白' or 桌腳=='mini三節黑' or 桌腳=='mini二節黑'or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳'):
                            桌腳 = 桌腳+'(腳底座45公分)'
                            total=桌板price+製材所形狀price+桌腳price
        
                            print('和您報價')
                            print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            
                        elif 桌深<72 and (桌腳=='force' or 桌腳 == 'force桌腳' or 桌腳 == 'force四柱桌腳' or 桌腳 == 'force四柱白腳' or 桌腳 == 'force四柱黑腳'):
                            print('桌深無法安裝force!')
                        
                        else:
                            total=桌板price+製材所形狀price+桌腳price
    
                            print('和您報價')
                            print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            
                elif 木種 == '琥珀木':
                    try:
                        桌寬=float(input('請輸入桌寬:'))
                        桌深=float(input('請輸入桌深:'))
                        厚度=float(input('厚度請輸入4.5:'))
                        桌腳=input('請輸入桌腳:') 
                        製材所形狀=input('請輸入桌板形狀:')
             
                        桌板price = 查詢琥珀木價格(桌寬, 桌深)*查詢木種設定(木種)
                        桌腳price = 查詢桌腳價格(桌腳)
                        運費price=查詢製材所單購桌板運費(木種)
                        
                        if 製材所形狀 =='':
                            製材所形狀 = '四方全平'
                        製材所形狀price=查詢製材所形狀價格(製材所形狀)
                    except Exception as e:
                        print('輸入錯誤，原因:', e)
                    else:
                        if 桌腳 == '':
                            total=桌板price+製材所形狀price+運費price
    
                            print('和您報價')
                            print('(製材所)-單購桌板訂%3.0f*%3.0f*%1.1f%s(%s)=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            
                        elif 桌深 >= 81 and 厚度<=2.7:
                                print('無法製作!!桌深厚度81cm以上，厚度需為3.3以上')
    
                        elif 桌寬<110 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰' or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳') and 桌深<57.5:
                            桌腳=桌腳+'(短版+短側片+腳底座45公分)'
                            total=桌板price+製材所形狀price+桌腳price
    
                            print('和您報價')
                            print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
        
                        elif 桌寬<110 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime二節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰' or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳') and 桌深>=57.7:
                            if 57.5<=桌深<60:
                                桌腳=桌腳+'(短版+腳底座45公分)'
                                total=桌板price+製材所形狀price+桌腳price
    
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                                    
                            elif 60<=桌深<68:
                                桌腳=桌腳+'(短版+腳底座60公分)'
                                total=桌板price+製材所形狀price+桌腳price
    
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            else:
                                桌腳=桌腳+'(短版)'
                                total=桌板price+製材所形狀price+桌腳price
    
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                        
               
                        elif 桌深<68 and (桌腳=='prime三節' or 桌腳=='prime二節' or 桌腳=='prime三節黑' or 桌腳=='prime二節黑'or 桌腳=='prime三節白' or 桌腳=='prime二節白' or 桌腳=='prime三節灰'or 桌腳=='prime二節灰'or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳'):
                            if 57.5<=桌深<68:
                                桌腳=桌腳+'(腳底座60公分)'
                                total=桌板price+製材所形狀price+桌腳price
        
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            elif 桌深<57.5:
                                桌腳=桌腳+'(短側片+腳底座45公分)'
                                total=桌板price+製材所形狀price+桌腳price
        
                                print('和您報價')
                                print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                                print('\n''備註:')
                                print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                                print('2. 訂製約45工作天')
                                print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                                
                        elif 桌深<60 and (桌腳=='mini三節' or 桌腳=='mini二節' or 桌腳=='mini三節白' or 桌腳=='mini二節白' or 桌腳=='mini三節黑' or 桌腳=='mini二節黑'or 桌腳=='固定桌腳' or 桌腳 =='固定白腳' or 桌腳 == '固定黑腳'):
                            桌腳 = 桌腳+'(腳底座45公分)'
                            total=桌板price+製材所形狀price+桌腳price
        
                            print('和您報價')
                            print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                            
                        elif 桌深<72 and 桌腳=='force':
                            print('桌深無法安裝force!')
                        
                        else:
                            total=桌板price+製材所形狀price+桌腳price
    
                            print('和您報價')
                            print('(製材所)-訂%3.0f*%3.0f*%1.1f%s(%s)+%s=%5.0f' % (桌寬,桌深,厚度,木種,製材所形狀,桌腳,total))
                            print('\n''備註:')
                            print('1. 以上金額含運(不含宜花東地區)、不含安裝')
                            print('2. 訂製約45工作天')
                            print('3. 製程為工廠製作時間，不包含後續的配送與安裝，實際配送日以通知為主')
                else:
                    print('木種輸入錯誤!!')

        elif 報價 == '報價單':
            # === 0. 路徑設定 ===============================================
            base_dir = Path.cwd()
            TEMPLATE   = base_dir / "FUNTE電動升降桌報價單-空白報價單.xlsx"
            PRODUCT_DB = base_dir / "報價單產品.xlsx"
            
            # === 1. 讀產品資料表 ==========================================
            items_df = pd.read_excel(PRODUCT_DB, header=14)
            items_df.columns = items_df.columns.str.strip()
            items_df.set_index("Item No.", inplace=True)
            
            # === 2. 客戶資料 ==============================================
            customer = {
                "name": input("姓名／公司名／統編：").strip(),
                "tel":  input("電話：").strip(),
                "addr": input("地址：").strip()
            }
            
            # === 3. 多品項輸入 ============================================
            orders = []
            while True:
                prod = input("產品名稱（Enter 結束）：").strip()
                if not prod:
                    break
                if prod not in items_df.index:
                    print("⚠️ 找不到此產品，請重輸")
                    continue
                orders.append({"row": items_df.loc[prod]})
            
            if not orders:
                print("⚠️ 沒有輸入任何產品，程式結束")
                exit()
            
            # === 4. 讀範本，寫客戶資料 ===================================
            wb = openpyxl.load_workbook(TEMPLATE)
            ws = wb.active
            ws["A9"].value  = f"Messrs：{customer['name']}"
            ws["A10"].value = f"TEL：{customer['tel']}"
            ws["A11"].value = f"ADDR：{customer['addr']}"
            ws["C11"].value = f"DATE：{date.today():%Y/%m/%d}"
            
            # === 5. 插入空白列：把原本總計/備註往下推 =====================
            HEADER_ROW  = 15        # 表頭在第 15 列
            start_row   = HEADER_ROW + 1            # 第一筆資料 → 16
            extra_rows  = len(orders) - 1
            if extra_rows > 0:
                ws.insert_rows(
                    idx=start_row + 1, amount=extra_rows)
            
            # === 6. 寫產品列 & 圖片 =====================================
            for i, od in enumerate(orders):
                row = start_row + i
                r = od["row"]
            
                #ws.cell(row, 1, r["Item No."])
                ws.cell(row, 2, r["Description"])
                ws.cell(row, 3).value = f"=ROUND({r["Price (NTD)"]}, 0)" 
                ws.cell(row, 5, f"=C{row}*D{row}")             # E 欄小計
            
                # 插圖（路徑用 base_dir）
                        
                item_no = str(r.name).strip()        # 取得「大洞洞板」
                img_path = base_dir / "報價單產品圖檔" / f"{item_no}.JPG"
                if img_path.exists():
                    img = Image(str(img_path))
                    img.width, img.height = 120, 90
                    ws.column_dimensions["A"].width = 18  # 大約適合120px圖片
                    ws.add_image(img, f"A{row}")
                else:
                    print(f"⚠️ 找不到圖片：{img_path}")
            
            # === 7. 重寫總計列公式 =======================================
            total_row = start_row + len(orders)
            ws.cell(total_row, 1, "總計新台幣（含稅）")
            ws.cell(total_row, 5, f"=SUM(E{start_row}:E{total_row-1})")
            
            # === 8. 另存檔案 =============================================
            safe_name = customer["name"].replace(" ", "_")
            outfile = base_dir / f"FUNTE電動升降桌報價單_{safe_name}_{date.today():%Y%m%d}.xlsx"
            wb.save(outfile)
            
            print(f"\n✅ 報價單已產生：{outfile}")
            
        elif 報價 == '下單':            
            ACCESS_TOKEN = 'SHOPLINE_API_KEY'
            API_BASE_URL = 'https://open.shopline.io/v1'

            headers = {
                'Authorization': f'Bearer {ACCESS_TOKEN}',
                'Content-Type': 'application/json'
            }

            try:
                custom_name = input('請輸入姓名: ')
                custom_product = input('請輸入商品名稱: ')
                custom_price = int(input('請輸入價格: '))
                custom_sku = input('請輸入物料(CD01/CD02):')
                #custom_url_code = input('請輸入連結編碼: ')
            except Exception as e:
                print('輸入錯誤，原因:', e)
            else:
                # 自動產生亂碼連結（14碼）
                custom_url_code = ''.join(random.choices(string.ascii_lowercase + string.digits, k=14))
                payload = {
                    "product": {
                        "title_translations": {
                            "zh-hant": f"【{custom_name}】-{custom_product}"
                        },
                        "category_ids":[
                            "5f3e6edecbccfd003c3fc43f","6620f0d9de0f7400174633cd"
                        ],
                        "summary_translations": {
                            "zh-hant": f"🔺製作天數不含配送時間，商品完成後將依訂單順序安排出貨"
                        },
                        "price": 99999,  # 原價，可視情況調整
                        "price_sale": custom_price,
                        "unlimited_quantity": True,
                        "sku": custom_sku,
                        "weight": 50,
                        "status": "hidden",
                        "blacklisted_payment_ids": [
                            "59d302ce080f065a31000066"
                        ],
                        "images": [
                            "SHOPLINE_IMAGES"
                        ],
                        "link": custom_url_code
                    }
                }

                product_url = f'{API_BASE_URL}/products'

                response = requests.post(product_url, headers=headers, json=payload)

                if response.status_code == 201:
                    base_url = 'https://www.funtetw.com/products/'
                    print(f'✅ 已生成連結：{base_url}{custom_url_code}')

                else:
                    print(f'❌ 建立失敗，狀態碼：{response.status_code}')
                    print(response.text)

        else:
            print('選項輸入錯誤!!')
    
input('請按enter鍵結束')
    #again = input("要重新報價嗎？(y/n)：")  
    #if again.lower() != 'y':
            #break